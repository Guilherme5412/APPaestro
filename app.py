import streamlit as st
import pandas as pd
import io
import os
import re
from datetime import datetime
from attendance_parser import parse_html_content
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

st.set_page_config(
    page_title="Paestro",
    page_icon="📋",
    layout="wide"
)

def initialize_session_state():
    """Initialize session state variables if they don't exist"""
    if 'classes' not in st.session_state:
        st.session_state.classes = {}
    if 'selected_class' not in st.session_state:
        st.session_state.selected_class = None
    if 'students' not in st.session_state:
        st.session_state.students = []
    
    # Agora serão dicionários de dicionários
    # Exemplo: attendance_status["TURMA X"]["Aluno Y"] = "P"
    if 'attendance_status' not in st.session_state:
        st.session_state.attendance_status = {}
    if 'observations' not in st.session_state:
        st.session_state.observations = {}
    
    if 'file_uploaded' not in st.session_state:
        st.session_state.file_uploaded = False
    if 'html_content' not in st.session_state:
        st.session_state.html_content = None

def handle_file_upload():
    """Process uploaded HTML file and extract class information"""
    uploaded_file = st.file_uploader("Fazer upload do arquivo HTML", type=["html", "htm"])
    
    if uploaded_file is not None and not st.session_state.file_uploaded:
        # Read file content
        html_content = uploaded_file.read().decode('utf-8')
        st.session_state.html_content = html_content
        
        try:
            # Parse HTML to extract classes and students
            classes = parse_html_content(html_content)
            
            if classes:
                st.session_state.classes = classes
                st.session_state.file_uploaded = True
                st.success("Arquivo carregado e processado com sucesso!")
            else:
                st.error("Nenhuma turma encontrada no arquivo enviado.")
        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")
    
    return st.session_state.file_uploaded

def display_class_selection():
    """Display dropdown for class selection"""
    class_names = list(st.session_state.classes.keys())
    
    st.subheader("Turmas Disponíveis")
    selected_class = st.selectbox(
        "Selecione uma turma:",
        class_names,
        index=0,
        placeholder="Escolha uma turma..."
    )
    
    if selected_class and selected_class != st.session_state.selected_class:
        st.session_state.selected_class = selected_class
        st.session_state.students = st.session_state.classes[selected_class]
        
        # -----------------------------------------------------------------------
        # MUDANÇA 1: Em vez de sobrescrever attendance_status e observations,
        # criamos/atualizamos subdicionários para a turma selecionada.
        # -----------------------------------------------------------------------
        
        # Se essa turma ainda não existir no attendance_status, inicializamos
        if selected_class not in st.session_state.attendance_status:
            st.session_state.attendance_status[selected_class] = {}
        if selected_class not in st.session_state.observations:
            st.session_state.observations[selected_class] = {}
        
        # Garante que cada aluno da turma tenha uma entrada no dicionário
        for student in st.session_state.students:
            if student not in st.session_state.attendance_status[selected_class]:
                st.session_state.attendance_status[selected_class][student] = None
            if student not in st.session_state.observations[selected_class]:
                st.session_state.observations[selected_class][student] = ""
        
        st.success(f"Carregados {len(st.session_state.students)} alunos da turma {selected_class}")
        st.rerun()

def display_attendance_form():
    """Display attendance form for selected class"""
    if not st.session_state.selected_class or not st.session_state.students:
        st.info("Por favor, selecione uma turma para marcar presença.")
        return
    
    st.subheader(f"Lista de Presença: {st.session_state.selected_class}")
    st.write(f"Total de alunos: {len(st.session_state.students)}")
    
    # Create a form for attendance marking
    with st.form("attendance_form"):
        for i, student in enumerate(st.session_state.students):
            col1, col2, col3 = st.columns([3, 2, 5])
            
            with col1:
                st.write(f"{i+1}. {student}")
            
            with col2:
                attendance_options = ["P", "F", "FJ"]
                
                # -----------------------------------------------------------------------
                # MUDANÇA 2: Pegamos o status do subdicionário da turma selecionada
                # -----------------------------------------------------------------------
                current_status = st.session_state.attendance_status[st.session_state.selected_class].get(student, None)
                
                index_value = attendance_options.index(current_status) if current_status in attendance_options else 0
                attendance_key = f"{st.session_state.selected_class}_attendance_idx_{i}"
                
                status = st.radio(
                    "Status",
                    attendance_options,
                    key=attendance_key,
                    index=index_value,
                    horizontal=True,
                    label_visibility="collapsed"
                )
            
            with col3:
                observation_key = f"{st.session_state.selected_class}_observation_idx_{i}"
                current_observation = st.session_state.observations[st.session_state.selected_class].get(student, "")
                
                observation = st.text_input(
                    "Observação",
                    value=current_observation,
                    key=observation_key,
                    placeholder="Adicionar observação (opcional)",
                    label_visibility="collapsed"
                )
            
            st.divider()
        
        submit_button = st.form_submit_button("Salvar Presença")
        
        if submit_button:
            # Update session state with form values
            for i, student in enumerate(st.session_state.students):
                attendance_key = f"{st.session_state.selected_class}_attendance_idx_{i}"
                observation_key = f"{st.session_state.selected_class}_observation_idx_{i}"
                
                if attendance_key in st.session_state:
                    status_value = st.session_state[attendance_key]
                    # -----------------------------------------------------------------------
                    # MUDANÇA 2 (continuação): Salvar no subdicionário da turma
                    # -----------------------------------------------------------------------
                    st.session_state.attendance_status[st.session_state.selected_class][student] = status_value
                
                if observation_key in st.session_state:
                    st.session_state.observations[st.session_state.selected_class][student] = st.session_state[observation_key]
            
            st.success("Dados de presença salvos com sucesso!")

def export_attendance():
    """Export attendance data as Excel (XLSX) with a header and group by turma"""
    if not st.session_state.file_uploaded or not st.session_state.classes:
        st.warning("Por favor, faça o upload de um arquivo HTML primeiro.")
        return
    
    st.subheader("Exportar Lista de Presença")
    
    # Tenta extrair o nome da escola do HTML, se possível
    school_name = "Escola"
    if st.session_state.html_content:
        match = re.search(r'(PREFEITURA MUNICIPAL [^\n]+)', st.session_state.html_content)
        if match:
            school_name = match.group(1).strip()
    
    # Cria um novo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Presença"
    
    # Cabeçalho com o nome da escola (mesclado em A1:C1)
    ws.merge_cells("A1:C1")
    ws["A1"] = school_name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    current_row = 3  # Inicia na linha 3
    
    # Itera por todas as turmas
    for turma, alunos in st.session_state.classes.items():
        # Adiciona a linha da turma
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        ws.cell(row=current_row, column=1).value = f"Turma: {turma}"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1
        
        # Cabeçalho da tabela para a turma
        ws.cell(row=current_row, column=1).value = "Aluno"
        ws.cell(row=current_row, column=2).value = "Presença"
        ws.cell(row=current_row, column=3).value = "Observação"
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).font = Font(bold=True)
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")
        current_row += 1
        
        # Para cada aluno da turma, adiciona os dados de presença e observação
        for student in alunos:
            ws.cell(row=current_row, column=1).value = student
            
            # -----------------------------------------------------------------------
            # MUDANÇA 3: Agora buscamos do subdicionário: st.session_state.attendance_status[turma]
            # -----------------------------------------------------------------------
            presence = ""
            observation = ""
            
            if turma in st.session_state.attendance_status:
                presence = st.session_state.attendance_status[turma].get(student, "")
            if turma in st.session_state.observations:
                observation = st.session_state.observations[turma].get(student, "")
            
            ws.cell(row=current_row, column=2).value = presence
            ws.cell(row=current_row, column=3).value = observation
            current_row += 1
        
        # Linha vazia entre turmas
        current_row += 1
    
    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    
    output = io.BytesIO()
    try:
        wb.save(output)
        excel_data = output.getvalue()
        file_base_name = f"Presenca_{datetime.now().strftime('%Y%m%d')}"
        st.download_button(
            "Baixar Excel (XLSX)",
            excel_data,
            f"{file_base_name}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Erro ao criar arquivo Excel: {e}")
        st.info("Por favor, baixe o arquivo CSV como alternativa.")

def reset_app():
    """Reset app state"""
    if st.button("Reiniciar Aplicação"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

def main():
    """Main application function"""
    st.title("📋 Sistema de Controle de Presença")
    
    # Initialize session state
    initialize_session_state()
    
    # App layout using tabs
    tab1, tab2, tab3 = st.tabs(["Upload", "Marcar Presença", "Exportar"])
    
    with tab1:
        st.header("Importar Arquivo HTML")
        file_uploaded = handle_file_upload()
        
        if file_uploaded:
            display_class_selection()
    
    with tab2:
        st.header("Marcar Presença")
        if st.session_state.file_uploaded:
            display_attendance_form()
        else:
            st.info("Por favor, faça o upload de um arquivo HTML primeiro.")
    
    with tab3:
        st.header("Exportar Dados")
        if st.session_state.file_uploaded and st.session_state.classes:
            export_attendance()
        else:
            st.info("Por favor, faça o upload de um arquivo HTML e marque a presença primeiro.")
    
    # Add a reset button on the sidebar
    st.sidebar.header("Ações")
    reset_app()

if __name__ == "__main__":
    main()
